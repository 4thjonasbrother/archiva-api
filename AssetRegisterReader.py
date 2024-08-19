from openpyxl import load_workbook
from typing import List, Dict, Tuple
import json


class AssetRegister:
    def __init__(self, filename: str = "NA Asset Register 2023.xlsx"):
        self.register = load_workbook(filename, read_only=True)

    def get_assets(self):
        """
        Get all assets' data. \n
        Returns a list of dictionaries. Each dictionary will be in the following format:
        
        ``{
            "asset_number": "",
            "sap_number": "",
            "name": "",
            "present_location": "",
            "condition": ""
        }``
        """
        sheets = self.register.sheetnames
        results: List[Dict[str, str]] = [] # Will store the details of all the assets

        for sheet_name in sheets:
            current_sheet = self.register[sheet_name]

            if current_sheet["A12"].value and current_sheet["A12"].value.strip() == "#":
                titles_row = 12
            elif current_sheet["A13"].value and current_sheet["A13"].value.strip() == "#":
                titles_row = 13
            elif current_sheet["A14"].value and current_sheet["A14"].value.strip() == "#":
                titles_row = 14
            elif current_sheet["A15"].value and current_sheet["A15"].value.strip() == "#":
                titles_row = 15
            elif current_sheet["A16"].value and current_sheet["A16"].value.strip() == "#":
                titles_row = 16

            titles: List[Tuple[str, str]] = []
            for row in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
                title: str | None = current_sheet[f"{row}{titles_row}"].value
                if title:
                    titles.append((title.lower().strip(), row))

            start_row = titles_row + 1

            for row in range(start_row, 100):
                data = {
                    "asset_number": "",
                    "sap_number": "",
                    "name": "",
                    "present_location": "",
                    "condition": ""
                }
                for each_title, column  in titles:
                    if each_title == "asset number":
                        asset_num_cell: str = current_sheet[f"{column}{row}"].value
                        if asset_num_cell:
                            data["asset_number"] = asset_num_cell.strip()

                    elif each_title in ["sap asset number", "sap asset no."]:
                        sap_num_cell: str | int = current_sheet[f"{column}{row}"].value
                        if sap_num_cell:
                            data["sap_number"] = str(sap_num_cell).strip()

                    elif "name" in each_title:
                        name_cell: str = current_sheet[f"{column}{row}"].value
                        if name_cell:
                            data["name"] = name_cell.strip()

                    elif "previous" in each_title:
                        previous_loc_cell: str = current_sheet[f"{column}{row}"].value
                        if previous_loc_cell:
                            data["previous_location"] = previous_loc_cell.strip()

                    elif "condition" in each_title:
                        condition_cell: str = current_sheet[f"{column}{row}"].value
                        if condition_cell:
                            data["condition"] = condition_cell.strip()

                if data != {
                    "asset_number": "",
                    "sap_number": "",
                    "name": "",
                    "previous_location": "",
                    "condition": ""
                }:
                    results.append(data)
                else:
                    break
        
        return results

    def export_to_json(self):
        """Exports the assets' information into a JSON file called `assets.json`"""
        data = self.get_assets()
        
        with open("assets.json", "w", encoding="utf-8") as json_file:
            json.dump(data, json_file, indent=4)
        
        return
    
    def read_assets_json_file(self):
        """Reads the data from the exported `assets.json` file"""
        try:
            with open("assets.json", "r", encoding="utf-8") as json_file:
                data: List[Dict[str, str]] = json.load(json_file)
            return data
        except FileNotFoundError:
            return []
    
if __name__ == "__main__":
    ar = AssetRegister()
    ar.read_assets_json_file()

            